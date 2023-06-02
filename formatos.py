# -*- coding: utf-8 -*-

import csv
import openpyxl
from datetime import datetime


def actualizar_archivo_excel(archivo_excel, archivo_csv, columnas, estados_validos: dict[str, str], tipos_incidencia_validos):
    """

    Args:
        archivo_excel (str): Ruta del archivo Excel a actualizar.
        archivo_csv (str): Ruta del archivo CSV a crear.
        columnas (dict): Diccionario que mapea las columnas de interés en el archivo Excel con los nombres correspondientes en el archivo CSV.
        estados_validos (dict): Diccionario que mapea los estados válidos en el archivo Excel a los estados correspondientes en el archivo CSV.
        tipos_incidencia_validos (list): Lista de tipos de incidencia válidos en el archivo Excel.
    """

    # Función para formatear la fecha
def formatear_fecha_fin(fecha_fin):
    if isinstance(fecha_fin, datetime):
        return fecha_fin.strftime('%d-%m-%Y')
    return fecha_fin


    # Función para formatear el estado


def formatear_estado(estado, estados_validos: dict[str, str]):
    estado_formateado = estados_validos.get(estado)

    # Función para formatear el tipo de incidencia

    def formatear_tipo_incidencia(tipo_incidencia,  tipos_incidencia_validos):
        print("tipo_incidencia")
        print(tipo_incidencia)
        if tipo_incidencia not in tipos_incidencia_validos:
            return "necesita mapeo"
        else:
            return tipos_incidencia_validos[tipo_incidencia]

    try:
        # Validar la existencia de los archivos
        if not (archivo_excel and archivo_csv):
            raise ValueError(
                "Debe proporcionar las rutas de archivo válidas.")
            
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(archivo_excel)
        
        # Seleccionar la hoja de trabajo
        worksheet = workbook.active
        
        # Buscar las cabeceras de las columnas
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
        header = [cell.value for cell in header_row]
        
        # Verificar que las columnas de interés estén presentes en el archivo Excel
        for nombre_columna in columnas.values():
            if nombre_columna not in header:
                raise ValueError(f'Cabecera no encontrada: {nombre_columna}')

        # Obtener los índices de las columnas de interés
        indice_columnas = {columna: header.index(
            nombre_columna) for columna, nombre_columna in columnas.items()}
        print("indices")
        print(indice_columnas)

        # Pedir al usuario la columna a modificar
        columnas_modificar = list(columnas.values())

        # Pedir al usuario la prioridad
        columnas_modificar = list(columnas.keys())
        
        # Crear archivo CSV
        with open(archivo_csv, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            # Escribir cabecera
            writer.writerow([columnas[columna] for columna in columnas])
            
            # Iterar sobre las filas y hacer los cambios necesarios
            for row in worksheet.iter_rows(min_row=2):
                
                # Inicializar una lista para almacenar los valores de las columnas
                fila_actualizada = []

                # Iterar sobre las columnas y aplicar los cambios necesarios
                for columna_modificar in columnas_modificar:

                    # Obtener el índice de la columna en el archivo Excel
                    indice_columna = indice_columnas.get(columna_modificar)

                    # Verificar si la columna existe en el archivo Excel
                    if indice_columna is None:
                        # Columna no encontrada, agregar un valor predeterminado en blanco
                        valor_actualizado = ""
                    else:
                        # Obtener el valor actualizado de la columna
                        print(valor_actualizado)

                        # Realizar el formateo correspondiente según la columna
                        if columna_modificar == 'fecha_fin':
                            valor_actualizado = formatear_fecha_fin(valor_actualizado)
                        elif columna_modificar == 'estado':
                            valor_actualizado = formatear_estado(
                                valor_actualizado, estados_validos)
                        elif columna_modificar == 'tipo_incidencia':
                            valor_actualizado = formatear_tipo_incidencia(
                                valor_actualizado,  tipos_incidencia_validos)

                    # Agregar el valor actualizado a la lista de la fila actualizada
                    fila_actualizada.append(valor_actualizado)

                # Escribir fila actualizada en archivo CSV
                writer.writerow(fila_actualizada)

        workbook.save(archivo_excel)
        workbook.close()

        print(f"Archivo {archivo_csv} actualizado con éxito!")

    except FileNotFoundError:
        print(f"Error: El archivo {archivo_excel} no existe.")
    except ValueError as error:
        print(f"Error: {error}")


# Pedir al usuario las rutas de los archivos y las columnas de interés
archivo_excel = input("Ingresa la ruta del archivo Excel a actualizar: ")
archivo_csv = input("Ingresa la ruta del archivo CSV a crear: ")
columnas = {
    'estado': 'Estado',
    'tipo_incidencia': 'Tipo de Incidencia',
    'fecha_fin': 'Fecha Fin',
    'prioridad': 'Prioridad'
}
estados_validos = {
    'EN CURSO': 'En progreso',
    'CERRADO': 'Cerrada',
    'PENDIENTE': 'Abierta'
}
tipos_incidencia_validos = {
    'Tarea Planificada': "tarea", 'Tarea no Planificada': "subtarea"}

actualizar_archivo_excel(archivo_excel, archivo_csv,
                        columnas, estados_validos, tipos_incidencia_validos)
