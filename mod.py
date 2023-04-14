import argparse
import openpyxl


# Definir argumentos de línea de comandos
parser = argparse.ArgumentParser(description='Actualizar archivo Excel')
parser.add_argument('-i', '--input', type=str, help='Archivo de entrada (formato .xlsx)', required=True)
parser.add_argument('-o', '--output', type=str, help='Archivo de salida (formato .xlsx)', required=True)
args = parser.parse_args()

# Configuración

archivoExcel = args.input
archivoExcelActualizado = args.output
columnaFechaVencimiento = 'Fecha de vencimiento'
columnaEstado = 'Estado'
valorEstadoProduccion = 'PRODUCCION'
valorEstadoCerrado = 'CERRADO'
columnaTipoIncidencia = 'Tipo de Incidencia'
tiposIncidencia = [
    'Funcionalidad Planificada',
    'Tarea Planificada',
    'Tarea No Planificada'
]
columnaCreada = 'Creada'

# Función para formatear la fecha
def formatearFecha(fecha):
    return fecha.strftime('%d%m%Y')

# Función para formatear el estado de la tarea
def formatearEstado(estado):
    if estado == valorEstadoProduccion:
        return valorEstadoCerrado
    else:
        return estado

# Función para formatear el tipo de incidencia
def formatearTipoIncidencia(tipoIncidencia):
    if tipoIncidencia not in tiposIncidencia:
        raise ValueError(f"Tipo de incidencia no válida: {tipoIncidencia}")
    else:
        return tipoIncidencia

# Función para actualizar una fila de datos
def actualizarFila(fila):
    # Cambiar el formato de la fecha de vencimiento
    if columnaFechaVencimiento in worksheet:
        fechaVencimiento = fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value
        fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value = formatearFecha(fechaVencimiento)

    # Cambiar el formato de la fecha de creación
    if columnaCreada in worksheet:
        fechaCreada = fila[worksheet[columnaCreada].column_letter + str(fila.row)].value
        fila[worksheet[columnaCreada].column_letter + str(fila.row)].value = formatearFecha(fechaCreada)

    # Cambiar el estado de la tarea
    if columnaEstado in worksheet:
        estado = fila[worksheet[columnaEstado].column_letter + str(fila.row)].value
        if estado == valorEstadoProduccion:
            fila[worksheet[columnaEstado].column_letter + str(fila.row)].value = valorEstadoCerrado

      # Cambiar el tipo de incidencia si es válido
    if columnaTipoIncidencia in worksheet:
        tipoIncidencia = fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value
        try:
            tipoIncidenciaFormateada = formatearTipoIncidencia(tipoIncidencia)
        except ValueError as e:
            print(f"Error en la fila {fila.row}: {str(e)}")
            return fila
    fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value = tipoIncidenciaFormateada

    return fila



try:
    # Leer el archivo Excel
    workbook = openpyxl.load_workbook(archivoExcel)

    # Seleccionar la hoja de trabajos
    sheetName = workbook.sheetnames[0]
    worksheet = workbook[sheetName]

    # Iterar sobre las filas y hacer los cambios necesarios
    for row in worksheet.iter_rows(min_row=2):
        actualizarFila(row)

    # Guardar el archivo Excel actualizado
    workbook.save(archivoExcelActualizado)
    print(f"Archivo {archivoExcelActualizado} actualizado con éxito!")


except Exception as e:
    print(f"Error al actualizar el archivo Excel: {str(e)}")