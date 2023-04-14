import openpyxl

# Configuración
archivoExcel = "C:\Work\jira-search-0b651e7f-5ce9-4912-9e53-c9a4cbd42e77.xlsx"
columnaFechaVencimiento = 'Fecha de vencimiento'
columnaEstado = 'Estado'
valorEstadoProduccion = 'PRODUCCION'
valorEstadoCerrado = 'CERRADO'
columnaTipoIncidencia = 'Tipo de Incidencia'
tiposIncidencia = [
    'Funcionalidad Planificada',
    'Tarea Planificada',
    'Tarea No Planificada'
]

# Función para formatear la fecha
def formatearFecha(fecha):
    return fecha.strftime('%d%m%Y')

# Función para actualizar una fila de datos
def actualizarFila(fila):
    # Cambiar el formato de la fecha
    fechaActualizada = fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value
    fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value = formatearFecha(fechaActualizada)

    # Cambiar el estado de la tarea
    estado = fila[worksheet[columnaEstado].column_letter + str(fila.row)].value
    if estado == valorEstadoProduccion:
        fila[worksheet[columnaEstado].column_letter + str(fila.row)].value = valorEstadoCerrado

    # Cambiar el tipo de incidencia si es válido
    tipoIncidencia = fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value
    if tipoIncidencia not in tiposIncidencia:
        print(f"Error: Tipo de incidencia no válida para la fila {fila.row}")
    else:
        fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value = tipoIncidencia

    return fila

try:
    # Leer el archivo Excel
    workbook = openpyxl.load_workbook(archivoExcel)

    # Seleccionar la hoja de trabajo
    sheetName = workbook.sheetnames[0]
    worksheet = workbook[sheetName]

    # Iterar sobre las filas y hacer los cambios necesarios
    for row in worksheet.iter_rows(min_row=2):
        actualizarFila(row)

    # Guardar el archivo Excel actualizado
    newArchivoExcel = archivoExcel.replace('.xlsx', "C:\Work\jira-search-0b651e7f-5ce9-4912-9e53-c9a4cbd42e77.xlsx")
    workbook.save(newArchivoExcel)

    print(f"Archivo {newArchivoExcel} actualizado con éxito!")

except Exception as e:
    print(f"Error al actualizar el archivo Excel: {str(e)}")
