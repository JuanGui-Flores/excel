import argparse
import openpyxl

# OBSERVACIONES:
# El archivo de salida siempre tiene que ser .csv no .xlsx. El de entrada no importa realmente, jira puede exportar como sea 🗸
# Poder escribir un comando para pasar por argumento el archivo de entrada y el de salida tipo: fechas.py -i archivo.xlsx -o archivo.csv 🗸

# TODO:
# Refactorizar la funcion actualizarFila para que sea mas escalable separando el codigo en funciones. (es probable que necesitemos actualizar mas columnas en el futuro) 
# Añadir excepciones especificas en caso de necesitar instrucciones para esa excepcion en concreto ej: FileNotFoundError(asegurese de que el archivo existe),  etc
# Añadir formato de fecha de columna "Creada"
# Añadir opcion de ignorar las columnas no existentes ej si no existe la columna de tipo de incidencia, que no de error y que siga con el resto de columnas


# Definir argumentos de línea de comandos
parser = argparse.ArgumentParser(description='Actualizar archivo Excel')
parser.add_argument('-i', '--input', type=str, help='Archivo de entrada (formato .xlsx)', required=True)
parser.add_argument('-o', '--output', type=str, help='Archivo de salida (formato .csv)', required=True)
args = parser.parse_args()

# Configuración

archivoExcel = args.input
archivoCSV = args.output
columnaFechaVencimiento = 'Fecha de vencimiento'
columnaEstado = 'Estado'
valorEstadoProduccion = 'PRODUCCION'
valorEstadoCerrado = 'CERRADO'
columnaTipoIncidencia = 'Tipo de Incidencia'
tiposIncidencia = [
    'Funcionalidad Planificada',
    'Tarea Planificada',
    'Tarea No Planificada'
]

# Función para formatear la fecha
def formatearFecha(fecha):
    return fecha.strftime('%d%m%Y')

# Función para actualizar una fila de datos
def actualizarFila(fila):


    # Cambiar el formato de la fecha
    fechaActualizada = fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value
    fila[worksheet[columnaFechaVencimiento].column_letter + str(fila.row)].value = formatearFecha(fechaActualizada)

    # Cambiar el estado de la tarea
    estado = fila[worksheet[columnaEstado].column_letter + str(fila.row)].value
    if estado == valorEstadoProduccion:
        fila[worksheet[columnaEstado].column_letter + str(fila.row)].value = valorEstadoCerrado

    # Cambiar el tipo de incidencia si es válido
    tipoIncidencia = fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value
    if tipoIncidencia not in tiposIncidencia:
        print(f"Error: Tipo de incidencia no válida para la fila {fila.row}")
    else:
        fila[worksheet[columnaTipoIncidencia].column_letter + str(fila.row)].value = tipoIncidencia

    return fila

try:
    # Leer el archivo Excel
    workbook = openpyxl.load_workbook(archivoExcel)

    # Seleccionar la hoja de trabajo
    sheetName = workbook.sheetnames[0]
    worksheet = workbook[sheetName]

    # Iterar sobre las filas y hacer los cambios necesarios
    for row in worksheet.iter_rows(min_row=2):
        actualizarFila(row)

    # Guardar el archivo Excel actualizado
    workbook.save(archivoCSV)
    print(f"Archivo {archivoCSV} actualizado con éxito!")

except Exception as e: # Muy buena practica el manejo de excepciones
    print(f"Error al actualizar el archivo Excel: {str(e)}")